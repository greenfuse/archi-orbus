import xml
import re
import os
from tkinter import filedialog

import openpyxl

homedir = os.path.expanduser('~')
filetypes = (
    ('xml files', '*.xml'),
    )
selectfile = filedialog.askopenfile(
    title="Select the Model Exchange file",
    initialdir=homedir,
    filetypes=filetypes
    )

if not selectfile:
    print("Nothing selected. Bye bye")
    exit()
    
filepath = selectfile.name
dirname, basename = os.path.split(filepath)
filename, filetype = os.path.splitext(basename)
destination_filepath = os.path.join(dirname, (filename + '.xlsx'))

iterparse =  xml.etree.ElementTree.iterparse(filepath, ('start', 'start-ns'))
events = "start", "start-ns"
ns = {}
for event, elem in xml.etree.ElementTree.iterparse(filepath, events):
    if event == "start-ns":
        if elem[0] == "":
            ns['xmlns'] = elem[1]
        else:
            ns[elem[0]] = elem[1]
        
xmlns = ns['xmlns']
xsi = '{' +  ns['xsi'] + "}"

if not 'www.opengroup.org/xsd/archimate' in xmlns:
    print("This is not a Model Exchange file")
    exit()

lst_exceptions = ["AndJunction", "OrJunction"]
dict_elements = {}
lst_elements = [["Type", "Name", "Identifier"]]
lst_relationships = [
    (
    "Lead: Name", 
    "Lead: Type", 
    "Relationship: Type", 
    "Member: Name",
    "Member: Type"
    )
    ]

tree = xml.etree.ElementTree.parse(filepath)
root = tree.getroot()
relationships = root.findall('xmlns:relationships/xmlns:relationship', ns)
elements = root.findall('xmlns:elements/xmlns:element', ns)
prop_defs = root.findall('xmlns:propertyDefinitions/xmlns:propertyDefinition', ns)
prop_dict = {}

if prop_defs:
    for prop_def in prop_defs:
        prop_name = prop_def.find('xmlns:name', ns).text
        if prop_name:
            prop_id = prop_def.attrib['identifier']
            prop_dict[prop_id] = prop_name
            lst_elements[0].append(prop_name)

for element in elements:
    identifier = element.attrib['identifier']
    element_type = element.attrib[xsi + "type"]
    if element_type not in lst_exceptions:
        # separate capitalised words
        orbus_element_type = re.sub( r"([A-Z])", r" \1", element_type).lstrip()
        orbus_element_type = orbus_element_type.capitalize()
        
        name = element.find('xmlns:name', ns).text
        element_details = [orbus_element_type, name, identifier]
        properties = element.findall('xmlns:properties/xmlns:property', ns)
        dict_prop = {}
        if properties:
            for property in properties:
                prop_ref = property.attrib['propertyDefinitionRef']
                prop_name = prop_dict[prop_ref]
                prop_value = property.find('xmlns:value', ns).text
                dict_prop[prop_name] = prop_value
            for prop_header in lst_elements[0][3:]:
                if prop_header in dict_prop:
                    element_details.append(dict_prop[prop_header])
                else:
                    element_details.append("")
        lst_elements.append(element_details)

for relationship in relationships:
    source_id = relationship.attrib['source']
    target_id = relationship.attrib['target']
    for item in lst_elements:
        if source_id == item[2]:
            source_type = item[0]
            source_name = item[1]
        if target_id == item[2]:
            target_type = item[0]
            target_name = item[1]
    if source_name and target_name:
        relationship_type = relationship.attrib[xsi + "type"]
        orbus_relationship_type = "Archimate: " + relationship_type
        relationship_details = (source_name, source_type, orbus_relationship_type, target_name, target_type)
        lst_relationships.append(relationship_details)

wb = openpyxl.Workbook()
ws_objects = wb.create_sheet("Objects", 0)


for orbus_object in lst_elements:
    orbus_object.pop(2)
    ws_objects.append(orbus_object)
    
ws_relationships = wb.create_sheet("Relationships", 1)

for relationship in lst_relationships:
    ws_relationships.append(relationship)
    
wb.save(destination_filepath)
print("Saved file " + destination_filepath)    
